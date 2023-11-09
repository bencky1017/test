#!/usr/bin/env python
# -*- coding=utf-8 -*-
# xlsx文件处理类
# 主要使用库openpyxl

from re import match
from time import strftime

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.chart import BarChart3D, PieChart, Reference
from openpyxl.worksheet.worksheet import Worksheet


class Xlsx():
    def __init__(self):
        """ 通过实例化可以创建一个Excel工作簿，使用save属性时即可完成文件创建 """
        # 文件
        self.filename = ''
        """ 文件名（路径） """
        self.workbook = None
        """ 工作簿类型 """
        self.worksheet = Worksheet
        """ 工作表类型 """
        self.sheetlists = []
        """ 工作表标题列表 """
        self.sheetname = ''
        """ 新建工作表名 """
        self.new_sheet = None
        """ 新增的工作表 """
        self.delsheet = ''
        """ 删除的工作表 """
        # 单元格定位
        self.cell = ''
        """ 单元格 """
        self.get_cell_col = ''
        """ 获取的单元格列 """
        self.get_cell_row = ''
        """ 获取的单元格行 """
        self.cell_col = ''
        """ 单元格列 """
        self.cell_row = ''
        """ 单元格行 """
        # 单元格坐标
        self.get_axis_col = ''
        """ 获取的坐标横轴 """
        self.get_axis_row = ''
        """ 获取的坐标横轴 """
        self.axis_col = ''
        """ 坐标纵轴 """
        self.axis_row = ''
        """ 坐标横轴 """

    # 创建指定名称的工作表

    def create(self, filename: str) -> Workbook:
        """
        创建文件
        @filename：必须参数
        """
        self.workbook = Workbook()
        self.workbook.save(filename)

    @property
    def save(self):
        self.workbook.save(self.filename)

    # 读取文件
    def readbook(self, filename: str = None) -> list:
        """
        读取文件
        @filename：必须参数
        """
        self.filename = filename
        if (filename is None):
            raise ValueError('\033[31mreadbook() 缺少1个必要参数：filename (文件名.xlsx）\033[0m')
        elif ('.xlsx' not in filename):
            raise ValueError('\033[31mfilename 参数设置错误，仅接受 .xlsx 格式\033[0m')
        # 读取文件
        self.workbook = load_workbook(filename=filename)
        print(f'读取文件：{filename}')
        # 显示表名
        self.sheetlists = self.__get_sheet_names()
        print(f'工作表名：{self.sheetlists}')
        return self.sheetlists

    # 读取sheet
    def readsheet(self, sheetname: str = None, display: bool = False):
        """ 
        ---
        读取工作表
        @sheetname：工作表名
        @display：是否显示

        :param sheetname：工作表名为必要参数
        :param display：为 False 不显示，为 True 显示
        """
        if sheetname in self.sheetlists:
            # 获取工作表
            self.worksheet = self.__get_sheet_by_name(sheetname)
            # self.worksheet.active
            print(f'表名：{sheetname}，类名：{self.worksheet}')
        elif sheetname is None or sheetname is '':
            raise TypeError(f'\033[31m [Error] readsheet() 函数中的 "sheetname" 参数缺失\033[0m')
        else:
            raise TypeError(f'\033[31m [Error] sheetname 参数 "{sheetname}" 不存在，注意区分大小写 \033[0m')
        # 行列数
        print(f"读取表：{self.worksheet.max_row} 行，{self.worksheet.max_column} 列")
        if display:
            print(f'{"-"*20} 读取数据开始 {"-"*20}')
            self.__read_sheet_value(self.worksheet.rows, 1, 1)
            print(f'{"-"*20} 数据读取结束 {"-"*20}')
        return self.worksheet

    # 通过index读取sheet
    def readsheet_index(self, index: int, display: bool = False):
        """
        通过index读取sheet
        @index：需要读取的index
        @display：是否显示

        :param index：从0开始
        :param display：为 False 不显示，为 True 显示
        """
        if index >= len(self.sheetlists):
            raise ValueError(f'\033[31m[Error] readsheet_index 中的参数 {index} 超出范围上限 {len(self.sheetlists)-1}\033[0m')
        else:
            # return self.__get_sheet_by_index(index)
            self.sheetname = self.sheetlists[index]
            # 获取工作表
            self.worksheet = self.__get_sheet_by_index(index)
            # self.worksheet.active
            print(f'表名：{self.sheetname}，类名：{self.worksheet}')
            # 行列数
            print(f"读取表：{self.worksheet.max_row} 行，{self.worksheet.max_column} 列")
            if display:
                print(f'{"-"*20} 读取数据开始 {"-"*20}')
                self.__read_sheet_value(self.worksheet.rows, 1, 1)
                print(f'{"-"*20} 数据读取结束 {"-"*20}')
            return self.worksheet

    # 获取单元格数据
    def readvalue(self, sheet: Worksheet):
        """
        获取单元格数据

        :param sheet:需要读取的表数据
        """
        data = []
        for row in sheet.rows:
            rowlist = []
            for cell in row:
                cell: Cell
                rowlist.append(cell.value)
            data.append(rowlist)
            # print(rowlist)
        return data

    # 添加sheet
    def addsheet(self, sheetname: str = '', data: list = [], cover=False) -> Worksheet:
        """
        ---
        添加工作表
        @sheetname：工作表名称
        @data：添加的数据列表，为空则不创建
        @cover：覆盖原有数据

        :param sheetname：不填写时候，默认命名为：工作表%m%d_%H%M%S，如：工作表0508_152953
        :param data：传入添加的数据列表`list[list]`，如:[['ID', 'Name', 'Age', 'City'], [1, 'John', 25, 'New York'], [2, 'Tom', 22, 'Washington']]
        :param cover：是否覆盖原有数据，默认不覆盖，为 True 时，sheetname 不能为空
        :type cover：bool
        """
        if (sheetname is ''):
            self.sheetname = '工作表'+strftime('%m%d_%H%M%S')
        elif (sheetname in self.sheetlists):
            self.sheetname = sheetname
            # 覆盖
            if cover:
                print(f'工作表：“{sheetname}” 已经存在，已覆盖写入')
            else:
                print(f'工作表：“{sheetname}” 已经存在，未覆盖')
                return
        else:
            # 如果没有此表名，则重置cover
            print(f'没有此表名：“{sheetname}”')
            cover = False
            self.sheetname = sheetname
        # """
        # 判断是否覆盖
        # 不覆盖：创建新数据表，写入数据
        # 覆盖：读取已有数据表，追加数据
        # """
        if not cover:
            self.new_sheet = self.__create_sheet(self.sheetname)
        else:
            self.new_sheet = self.readsheet(sheetname)

        print(f'添加了：{self.sheetname}')
        print(f'添加后工作表名：{self.__get_sheet_names()}')

        # 插入数据为空
        if len(data) == 0:
            return self.new_sheet

        # 插入数据不为空，追加数据
        cover_flag = 2
        # 从尾部追加
        if cover_flag == 1:
            for row in data:
                self.new_sheet.append(row)
        # 从头部覆盖
        elif cover_flag == 2:
            for i in range(len(data)):
                for j in range(len(data[i])):
                    self.__to_cell(self.new_sheet, i+1, j+1, data[i][j])
                    # cell_ref = self.new_sheet.cell(row=i+2, column=j+2).coordinate
                    # self.new_sheet[cell_ref] = data[i][j]
        # TODO：指定开始单元格开始覆盖
        else:
            start_row = 1
            start_col = 4
            for i in range(len(data)):
                for j in range(len(data[i])):
                    cell_ref = self.new_sheet.cell(row=start_row+i, column=start_col+j).coordinate
                    self.new_sheet[cell_ref] = data[i][j]
        self.save
        # 显示数据
        # temp_list = self.__read_sheet_value(self.new_sheet.rows, 3)
        # [print(row) for row in temp_list]
        return self.new_sheet

    # 删除sheet
    def deletesheet(self, sheetname: str) -> bool:
        """
        删除指定名称的sheet，返回删除的状态值
        @sheetname：需要删除的sheet名

        返回值:
            bool:True    --> 200 --> 删除成功
            bool:False   --> 400 --> 未找到文件
        """
        self.delsheet = sheetname
        if (sheetname not in self.__get_sheet_names()):
            print(f'未找到名为 “{sheetname}” 的工作表')
            return False
        else:
            del self.workbook[sheetname]
            print(f'删除了：{self.delsheet}')
            self.sheetlists = self.__get_sheet_names()
            print(f'删除后工作表名：{self.sheetlists}')
            self.save
            return True

    # 通过index删除sheet
    def deletesheet_index(self, index: int) -> bool:
        """
        通过index删除指定的sheet，返回删除的状态值
        @index：需要删除的index

        返回值:
            bool:True    --> 200 --> 删除成功
            bool:False   --> 400 --> 未找到文件
        :param index：从0开始
        """
        if index >= len(self.sheetlists):
            raise ValueError(f'\033[31m[Error] deletesheet_index 中的参数 {index} 超出范围上限 {len(self.sheetlists)-1}\033[0m')
        else:
            self.delsheet = self.sheetlists[index]
            del self.workbook[self.delsheet]
            print(f'删除了：{self.delsheet}')
            self.sheetlists = self.__get_sheet_names()
            print(f'删除后工作表名：{self.sheetlists}')
            self.save
            return True

    # 写入数据_行列rank
    def edit(self, row: int, col: int, content: str, __call: bool = True):
        """
        （默认）在第 row 行，第 col 列，写入 content 数据
        @row：行
        @col：列
        @content：写入内容
        """
        # 坐标
        self.get_axis_col = col
        self.get_axis_row = row
        get_col = col
        axis_col = ''
        while get_col > 0:
            axis_col = chr(65+(get_col-1) % 26) + axis_col
            get_col = (get_col-1) // 26
        self.axis_col = axis_col
        self.axis_row = row
        print(f'单元格坐标：{self.axis_row} 行, {self.axis_col} 列 / {self.get_axis_col} 列') if __call else None
        # self.worksheet.cell(row, col).value = content
        self.worksheet.cell(row, col, content)
        self.save

    # 写入数据_单元格cell
    def edit_cell(self, cell: str, content: str):
        """
        在 cell 单元格写入 content 数据
        @cell：单元格
        @content：写入内容
        - 单元格形式，如：A3，C11，AB12
        """
        reg = '([A-Z]+)(\d+)'
        self.cell = cell
        # 解析单元格
        self.get_cell_col, self.get_cell_row = match(reg, cell.upper()).groups()
        # 列转换数字
        temp_col_list = list(self.get_cell_col)
        temp_col_list_len = len(temp_col_list)
        temp_index_list = list(range(temp_col_list_len))[::-1]
        sum_col = 0
        for index in range(temp_col_list_len):
            # print(f'当前：{index},{temp_col_list[index]}')
            sum_col += (ord(temp_col_list[index])-64)*26**(temp_index_list[index])
        self.cell_col = sum_col
        # 行转换数字
        self.cell_row = int(self.get_cell_row)
        print(f'单元格位置：{self.get_cell_col} 列 / {self.cell_col} 列, {self.cell_row} 行')
        # 写入数据
        self.edit(self.cell_row, self.cell_col, content, False)
        self.save

    # 插入图表  TODO
    def insert_pie(self, data_list):
        print('插入图片')
        data = [
            ['优先级', '缺陷数', '修复数', '修复率'],
            ['严重', 4, 3, '75.00%'],
            ['主要', 8, 7, '87.50%'],
            ['次要', 6, 5, '83.33%'],
            ['不重要', 7, 5, '71.43%'],
            ['无优先级', 20, 20, '100%'],
            ['修改时间', '05-29', '15:06:34']
        ]

        wb = Workbook()
        ws = wb.active

        for row in data:
            ws.append(row)

        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=2, max_row=5)
        data = Reference(ws, min_col=2, min_row=1, max_row=5)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "Pies sold by category"

        # Cut the first slice out of the pie
        # slice = DataPoint(idx=0, explosion=20)
        # pie.series[0].data_points = [slice]

        ws.add_chart(pie, "D1")

        wb.save("pie.xlsx")

    """ *********************************** ↓↓↓↓↓ 内部调用方法 ↓↓↓↓↓ *********************************** """

    # 创建sheet工作表
    def __create_sheet(self, sheetname) -> Worksheet:
        self.sheetlists.append(sheetname)
        return self.workbook.create_sheet(sheetname)  # 尾端插入

    # 获取sheet名称列表
    def __get_sheet_names(self):
        return self.workbook.sheetnames

    # 获取指定sheet
    def __get_sheet_by_name(self, sheetname):
        return self.workbook[sheetname]

    # 获取指定index的sheet
    def __get_sheet_by_index(self, index):
        return self.workbook.worksheets[index]

    # 编辑单元格
    def __edit_cell(self, cell, value):
        self.worksheet[cell] = value
        return self.worksheet

    # 行列转单元格
    def __to_cell(self, new_sheet: Worksheet, row: int, col: int, value=None):
        """
        行列转为单元格索引，如：1 行 A 列--> A1
        @row：行
        @col：列
        @value：传入的数据

        :param row：行，至少为1
        :param col：列，至少为1
        :param value：传入的数据
        """
        # **
        # 此处的new_sheet其他地方可以是self.worksheet的工作表类型
        # 临时修改，后续删除
        # *
        rank_to_cell = new_sheet.cell(row=row, column=col).coordinate
        new_sheet[rank_to_cell] = value

    # 读取并解析数据
    def __read_sheet_value(self, rows, flag, display=0):
        """
        读取表格中的数据，将Worksheet类型转为list
        @rows：数据的行数
        @flag：方法标记，分为1,2,3，3个方法标记

        :param rows:数据行，一般为 Worksheet.rows
        :param flag:方法标记，不同方法操作不同，默认为 1
        :param display:显示标记，默认不显示，为1时候显示
        """
        method_flag = flag  # 1,2,3
        if (method_flag == 1):
            # 方法一
            rows = list(rows)
            data = []
            for row in rows:
                rowlist = list([])
                for cell in row:
                    cell: Cell
                    rowlist.append(cell.value)
                row_data = [cell.value for cell in row]
                data.append(row_data)
                if display == 1 or display == 3:
                    print(rowlist)
            if display == 2 or display == 3:
                print(data)
        elif (method_flag == 2):
            # 方法二——不能打印过程
            data = []
            for row in rows:
                data.append([cell.value for cell in row])
            # print(data)
        else:
            # 方法三
            data = []
            for row in rows:
                rowlist = []
                for cell in row:
                    rowlist.append(cell.value)
                data.append(rowlist)
                if display == 1 or display == 3:
                    print(rowlist)
            if display == 2 or display == 3:
                print(data)
        return data
    """ ************************************ ↑↑↑↑↑ 内部调用方法 ↑↑↑↑↑ ************************************ """


def test():
    """
    测试Excel表格增删改查操作

    |状态|序号|效果|
    |:--:|:--:|:--|
    |创建工作簿|1|创建新文件，指定文件名|
    |读取工作簿|2|读取文件，需要事先运行创建方法|
    |添加工作表|3|从 A1 开始添加表格数据|
    |删除工作表|4|删除指定的工作表文件|
    |读取工作表|5|读取指定工作表内容|
    """

    # 创建xlsx文件✅
    xlsx = 'test.xlsx'
    table = Xlsx()          # 创建一个 Xlsx 实例
    table.create(xlsx)      # 使用创建方法创建一个 xlsx 文件
    print(f'\033[32m[INFO] 创建名为 “{xlsx}” 的文件\033[0m')

    # 读取xlsx文件✅
    table.readbook(xlsx)    # 读取 xlsx 工作簿
    print(f'\033[32m[INFO] 读取 “{xlsx}” 文件，其中工作表有：{table.sheetlists}\033[0m')

    # 添加工作表✅
    sheetname = '测试'
    table.addsheet(sheetname, [['ID', 'Name', 'Age', 'City'], [1, 'John', 25, 'New York'], [2, 'Tom', 22, 'Washington']])
    print(f'\033[32m[INFO] 添加名为 “{table.sheetname}” 的工作表，添加后的工作表：{table.sheetlists}\033[0m')

    # 删除工作表✅
    # status = table.deletesheet_index(0)
    status = table.deletesheet('Sheet')
    if status:
        print(f'\033[32m[INFO] 删除名为 “{table.delsheet}” 的工作表，删除后的工作表：{table.sheetlists}\033[0m')
    else:
        print(f'\033[31m[ERROR] 删除 “{table.delsheet}” 失败❌\033[0m')

    # 读取工作表✅
    # table.readsheet(sheetname,1)
    sheet = table.readsheet_index(0, 1)
    print(f'\033[32m[INFO] 读取 “{table.sheetname}” 工作表数据\033[0m')

    # 单元格操作
    # 方法一
    sheet['A4'] = 3
    table.save

    # 方法二
    table.edit(4, 2, '添加的数据')
    table.edit_cell('ab13', 'edit_cell添加')
    print(f'\033[32m[INFO] 编辑单元格内容 “{table.cell}” 工作表数据\033[0m')


# if __name__ == '__main__':
#     test()
